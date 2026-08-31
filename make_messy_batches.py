"""
Splits sales_figures into monthly export batches (simulating separate
raw exports arriving over time), and deliberately introduces messiness
into some of them -- this is the "before" state the cleaning macro
will fix.

Messiness introduced (only in batches flagged MESSY):
- duplicate rows (some rows repeated)
- inconsistent MonthYear format (e.g. "2016-10" vs "10.2016" vs "Oct-16")
- blank/invalid required fields (missing StoreID, missing Turnover)
- inconsistent text casing in Dept. Name / City ("dry", "DRY", "Dry ")
- stray whitespace in text fields
"""
import openpyxl
import random
import copy

random.seed(42)

SRC = "salesworkload_ORIGINAL.xlsx"
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb["sales_figures"]

header = [c.value for c in ws[2]]  # row 2 is the real header
data_rows = [list(r) for r in ws.iter_rows(min_row=3, values_only=True)]

# Exclude stray footer/artifact rows already present in the source file
# (8 rows with MonthYear == ' - - - - ' and all other fields blank --
# this is a pre-existing export artifact, not something we introduced)
before = len(data_rows)
data_rows = [r for r in data_rows if r[0] != " - - - - "]
print(f"Excluded {before - len(data_rows)} pre-existing footer/artifact rows")

print("Header:", header)
print("Total data rows:", len(data_rows))

# Group rows by MonthYear to make batches that mirror real "monthly exports"
from collections import defaultdict
by_month = defaultdict(list)
for r in data_rows:
    by_month[r[0]].append(r)

months = sorted(by_month.keys())
print("Months found:", months)

# Take first 4 months as our 4 "export batches"
batch_months = months[:4]
print("Using batches:", batch_months)

MESSY_BATCHES = {batch_months[1], batch_months[3]}  # mess up 2 of the 4

def messify_month_year(val, style):
    # val like "10.2016"
    mm, yyyy = val.split(".")
    if style == "iso":
        return f"{yyyy}-{mm}"
    if style == "monname":
        months_short = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
        return f"{months_short[int(mm)-1]}-{yyyy[2:]}"
    return val  # unchanged

def messify_case(val):
    if val is None:
        return val
    styles = [lambda s: s.upper(), lambda s: s.lower(), lambda s: f"  {s}  "]
    return random.choice(styles)(val)

for month in batch_months:
    rows = copy.deepcopy(by_month[month])
    is_messy = month in MESSY_BATCHES

    if is_messy:
        n = len(rows)
        # 1) duplicate ~3% of rows (append duplicates)
        dup_count = max(2, int(n * 0.03))
        dups = [copy.deepcopy(random.choice(rows)) for _ in range(dup_count)]
        rows.extend(dups)

        # 2) inconsistent MonthYear format on ~15% of rows
        idxs = random.sample(range(len(rows)), int(len(rows) * 0.15))
        for i in idxs:
            style = random.choice(["iso", "monname", "same"])
            rows[i][0] = messify_month_year(rows[i][0], style)

        # 3) blank/invalid required fields on ~5% of rows
        idxs = random.sample(range(len(rows)), int(len(rows) * 0.05))
        for i in idxs:
            field = random.choice(["StoreID", "Turnover"])
            col = header.index(field)
            rows[i][col] = None

        # 4) inconsistent casing/whitespace on Dept. Name and City ~20%
        idxs = random.sample(range(len(rows)), int(len(rows) * 0.20))
        dept_col = header.index("Dept. Name")
        city_col = header.index("City")
        for i in idxs:
            rows[i][dept_col] = messify_case(rows[i][dept_col])
            if random.random() < 0.5:
                rows[i][city_col] = messify_case(rows[i][city_col])

        random.shuffle(rows)  # so duplicates aren't all at the end (realistic)

    # Write out this batch as its own file
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Raw_Export"
    out_ws.append(header)
    for r in rows:
        out_ws.append(r)

    safe_month = month.replace(".", "_")
    tag = "MESSY" if is_messy else "clean"
    fname = f"export_batch_{safe_month}_{tag}.xlsx"
    out_wb.save(fname)
    print(f"Wrote {fname}: {len(rows)} rows (original {len(by_month[month])})")
