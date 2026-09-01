"""
Builds the workbook shell: creates all sheets the project needs (empty,
except Raw_Data which gets the combined 4 batches imported).

This is STRUCTURE ONLY. No VBA macros yet, no cleaning applied yet,
no lookups/pivots/formatting yet -- those are separate, later steps.
"""
import os
import openpyxl

# scripts/ -> project root, and project root -> source_data
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.normpath(os.path.join(SCRIPT_DIR, ".."))
SOURCE_DIR = os.path.join(ROOT_DIR, "source_data")

BATCH_FILES = [
    "export_batch_01_2017_clean.xlsx",
    "export_batch_02_2017_MESSY.xlsx",
    "export_batch_03_2017_clean.xlsx",
    "export_batch_04_2017_MESSY.xlsx",
]

OUT_NAME = "RetailSalesWorkloadAutomation.xlsx"
# NOTE: saved as .xlsx here (no VBA yet). We'll convert/resave as .xlsm
# once we actually add macro code -- Excel/openpyxl distinguish the two,
# and there's no VBA to embed yet.

def get_unique_path(directory, filename):
    """
    Returns a path for `filename` in `directory`. If a file with that name
    already exists, appends " (1)", " (2)", etc. before the extension until
    a free name is found -- e.g. RetailSalesWorkloadAutomation (1).xlsx
    """
    base, ext = os.path.splitext(filename)
    candidate = os.path.join(directory, filename)
    counter = 1
    while os.path.exists(candidate):
        candidate = os.path.join(directory, f"{base} ({counter}){ext}")
        counter += 1
    return candidate

OUT_FILE = get_unique_path(ROOT_DIR, OUT_NAME)

wb = openpyxl.Workbook()

# --- Raw_Data sheet: combined import of all 4 batches, messiness intact ---
raw = wb.active
raw.title = "Raw_Data"

header = None
combined_rows = []
for f in BATCH_FILES:
    src_path = os.path.join(SOURCE_DIR, f)
    src_wb = openpyxl.load_workbook(src_path)
    src_ws = src_wb.active
    rows = list(src_ws.iter_rows(values_only=True))
    if header is None:
        header = rows[0]
    combined_rows.extend(rows[1:])

raw.append(header)
for r in combined_rows:
    raw.append(r)

print(f"Raw_Data: {len(combined_rows)} rows imported (header + data)")

# --- Store_Master sheet: cleaned-up store list from opening_schemes ---
store_master = wb.create_sheet("Store_Master")
orig_path = os.path.join(SOURCE_DIR, "salesworkload_ORIGINAL.xlsx")
orig_wb = openpyxl.load_workbook(orig_path, data_only=True)
schemes_ws = orig_wb["opening_schemes"]

# Header row for opening_schemes is row 6 (id, Store name, Region, Scheme, months...)
store_master.append(["id", "Store name", "Region", "Scheme"])
for row in schemes_ws.iter_rows(min_row=7, values_only=True):
    if row[0] is None:
        continue
    store_master.append([row[0], row[1], row[2], row[3]])

print(f"Store_Master: {store_master.max_row - 1} stores imported")

# --- Remaining sheets: created empty, built out in later steps ---
for name in [
    "Clean_Data",
    "Data_Entry_Form",
    "Reports_Pivot",
    "Alerts",
    "Cleaning_Log",
    "Summary_Export",
]:
    wb.create_sheet(name)

wb.save(OUT_FILE)
print(f"Saved {OUT_FILE} with sheets: {wb.sheetnames}")
